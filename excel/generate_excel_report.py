"""
excel/generate_excel_report.py
-------------------------------
STEP 14 — Excel Report.

Purpose:
    Build a multi-sheet Excel workbook (olist_data_quality_report.xlsx)
    from the pipeline's JSON/CSV outputs, suitable for a business
    stakeholder who wants to review data quality and governance results
    outside of Power BI.

Sheets produced:
    1. Summary            — pipeline run KPIs
    2. Quality Metrics     — dimension/dataset quality scores
    3. Issue Register        — full issue register with conditional formatting
    4. Reconciliation          — row/key/monetary reconciliation checks
    5. Dataset Profile           — profiling stats per dataset

How to run:
    python -m excel.generate_excel_report
    (run this AFTER the main pipeline so reports/*.json and
    reports/issue_register.csv already exist)

Connects to:
    - reports/profile_report.json
    - reports/issue_register.csv
    - reports/quality_scorecard.json
    - reports/reconciliation_report.json
    - config.py -> EXCEL_REPORT_PATH
"""

from __future__ import annotations

import json

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from src.config import (
    EXCEL_REPORT_PATH,
    ISSUE_REGISTER_PATH,
    PIPELINE_RUN_REPORT_PATH,
    PROFILE_REPORT_PATH,
    QUALITY_SCORECARD_PATH,
    RECONCILIATION_REPORT_PATH,
)
from src.logging_setup import get_logger

logger = get_logger(__name__)

HEADER_FILL = PatternFill(start_color="1F2A44", end_color="1F2A44", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=14, bold=True, color="1F2A44")


def _style_header_row(ws, row_idx: int, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _autofit(ws, df: pd.DataFrame, start_col: int = 1) -> None:
    for i, col in enumerate(df.columns, start=start_col):
        max_len = df[col].astype(str).str.len().max()
        max_len = 10 if pd.isna(max_len) else int(max_len)
        width = max(12, min(45, max_len + 2))
        ws.column_dimensions[get_column_letter(i)].width = width


def _write_df_as_table(ws, df: pd.DataFrame, table_name: str, start_row: int = 1) -> None:
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=start_row, column=j, value=col)
    for i, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)

    _style_header_row(ws, start_row, len(df.columns))
    _autofit(ws, df)

    end_row = start_row + len(df)
    end_col = get_column_letter(len(df.columns))
    if len(df) > 0:
        ref = f"{get_column_letter(1)}{start_row}:{end_col}{end_row}"
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True
        )
        ws.add_table(table)


def build_summary_sheet(wb: Workbook, run_report: dict) -> None:
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Olist Data Management, Quality & Governance — Pipeline Summary"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    rows = [
        ("Pipeline Run Timestamp (UTC)", run_report.get("pipeline_run_timestamp")),
        ("Pipeline Status", run_report.get("pipeline_status")),
        ("Total Records Processed", run_report.get("total_records_processed")),
        ("Overall Quality Score (%)", run_report.get("quality_scorecard", {}).get("overall_score_pct")),
        ("Reconciliation Status", run_report.get("reconciliation_summary", {}).get("overall_status")),
        ("Open Issues", run_report.get("validation_summary", {}).get("open_issues")),
        ("Critical Open Issues", run_report.get("validation_summary", {}).get("critical_open_issues")),
    ]
    for i, (label, value) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 30


def build_quality_metrics_sheet(wb: Workbook, scorecard: dict) -> None:
    ws = wb.create_sheet("Quality Metrics")
    rows = []
    for dataset, info in scorecard.get("dataset_scores", {}).items():
        for dim, score in info.get("dimension_scores_pct", {}).items():
            rows.append({"dataset": dataset, "dimension": dim, "score_pct": score})
    df = pd.DataFrame(rows) if rows else pd.DataFrame(columns=["dataset", "dimension", "score_pct"])
    _write_df_as_table(ws, df, "QualityMetricsTable")

    if not df.empty:
        chart = BarChart()
        chart.title = "Quality Score by Dataset & Dimension (%)"
        chart.y_axis.title = "Score (%)"
        data = Reference(ws, min_col=3, min_row=1, max_row=len(df) + 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(df) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, f"F2")


def build_issue_register_sheet(wb: Workbook, issue_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Issue Register")
    _write_df_as_table(ws, issue_df, "IssueRegisterTable")

    if not issue_df.empty and "severity" in issue_df.columns:
        severity_col_idx = list(issue_df.columns).index("severity") + 1
        col_letter = get_column_letter(severity_col_idx)
        rng = f"{col_letter}2:{col_letter}{len(issue_df) + 1}"
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator="equal", formula=['"Critical"'],
                       fill=PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")),
        )
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator="equal", formula=['"High"'],
                       fill=PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")),
        )


def build_reconciliation_sheet(wb: Workbook, reconciliation: dict) -> None:
    ws = wb.create_sheet("Reconciliation")
    all_checks = (
        reconciliation.get("row_count_checks", [])
        + reconciliation.get("key_count_checks", [])
        + reconciliation.get("monetary_total_checks", [])
    )
    df = pd.DataFrame(all_checks) if all_checks else pd.DataFrame(columns=["check", "dataset", "status"])
    _write_df_as_table(ws, df, "ReconciliationTable")

    if not df.empty and "status" in df.columns:
        status_col_idx = list(df.columns).index("status") + 1
        col_letter = get_column_letter(status_col_idx)
        rng = f"{col_letter}2:{col_letter}{len(df) + 1}"
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="equal", formula=['"FAIL"'],
                             fill=PatternFill(start_color="FF8080", end_color="FF8080", fill_type="solid")),
        )
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="equal", formula=['"WARNING"'],
                             fill=PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")),
        )
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="equal", formula=['"PASS"'],
                             fill=PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")),
        )


def build_dataset_profile_sheet(wb: Workbook, profile: dict) -> None:
    ws = wb.create_sheet("Dataset Profile")
    rows = []
    for name, d in profile.get("datasets", {}).items():
        rows.append({
            "dataset": name,
            "row_count": d.get("row_count"),
            "column_count": d.get("column_count"),
            "missing_cells_total": d.get("missing_cells_total"),
            "duplicate_full_rows": d.get("duplicate_full_rows"),
            "primary_key_candidates": ", ".join(d.get("primary_key_candidates", [])),
        })
    df = pd.DataFrame(rows)
    _write_df_as_table(ws, df, "DatasetProfileTable")


def generate_excel_report() -> None:
    with open(PROFILE_REPORT_PATH) as f:
        profile = json.load(f)
    with open(QUALITY_SCORECARD_PATH) as f:
        scorecard = json.load(f)
    with open(RECONCILIATION_REPORT_PATH) as f:
        reconciliation = json.load(f)
    with open(PIPELINE_RUN_REPORT_PATH) as f:
        run_report = json.load(f)
    issue_df = pd.read_csv(ISSUE_REGISTER_PATH)

    wb = Workbook()
    build_summary_sheet(wb, run_report)
    build_quality_metrics_sheet(wb, scorecard)
    build_issue_register_sheet(wb, issue_df)
    build_reconciliation_sheet(wb, reconciliation)
    build_dataset_profile_sheet(wb, profile)

    EXCEL_REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(EXCEL_REPORT_PATH)
    logger.info("Excel report generated at %s", EXCEL_REPORT_PATH)


if __name__ == "__main__":
    generate_excel_report()
